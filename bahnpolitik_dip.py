#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bahnpolitik-Tracker ueber die DIP-API des Deutschen Bundestags.
Gleiche Machart wie der MEX13-Sammler: laeuft in Pydroid oder GitHub Actions.

Version 3 - neu gegenueber Version 2:
  * Abstimmungsverhalten: ueber den Endpunkt /vorgangsposition wird zu jeder
    Drucksache der zugehoerige Vorgang geladen und daraus die
    Beschlussfassung ausgelesen (Tenor, Abstimmungsart, Mehrheit und die
    Bemerkung mit dem Fraktionsverhalten).
  * Die Bemerkung wird geparst: je Fraktion wird Zustimmung, Ablehnung oder
    Enthaltung ermittelt und in der Tabelle abstimmungen gespeichert.
  * Neues Excel-Blatt "Abstimmungen": farbige Matrix Vorlage x Fraktion,
    darunter eine Bilanz je Fraktion mit gestapeltem Balkendiagramm.
  * Neue Spalte "Antwort in" auf dem Blatt Drucksachen (Gegenrichtung zu
    "Bezug auf"): zeigt, ob und mit welcher Drucksache eine Kleine Anfrage
    beantwortet wurde. Leere Zelle = keine Antwort im Bestand.

Wichtig zur Erwartungshaltung: Ueber Kleine Anfragen und deren Antworten wird
nie abgestimmt. Abstimmungsfaehig sind nur Antraege, Gesetzentwuerfe und
Entschliessungsantraege - das ist ein kleiner Teil des Bestands. Findet das
Skript keine Beschlussfassungen, bleibt das Blatt "Abstimmungen" leer und
sagt das auch.

Version 2 - Aenderungen gegenueber Version 1:
  * Urheber-Zuordnung korrigiert: Gemeinschafts-Drucksachen mehrerer Fraktionen
    werden nicht mehr willkuerlich einer einzigen Fraktion zugeschlagen,
    sondern als "Mehrere Fraktionen" gefuehrt.
  * Neue Spalte "Rolle": Regierungsfraktion / Opposition je Wahlperiode.
    Ohne diese Spalte ist jeder Fraktionsvergleich irrefuehrend, weil
    Regierungsfraktionen praktisch keine Kleinen Anfragen stellen.
  * Neue Spalte "Themen": Mehrfach-Verschlagwortung aus dem Titel.
  * Antwortzeiten: Antworten der Bundesregierung werden ueber die im Titel
    genannte Drucksachennummer der zugehoerigen Kleinen Anfrage zugeordnet,
    die Bearbeitungsdauer in Tagen wird berechnet.
  * Excel komplett neu: Dashboard mit Kennzahlen, Initiativen je Fraktion
    (nur eigeninitiative Drucksachentypen), Zeitverlauf nach Quartalen,
    Themenmatrix, Antwortzeiten-Auswertung. Alle Auswertungsblaetter
    arbeiten mit Formeln, rechnen sich also nach Filtern neu.

Was das Skript tut:
  1. Sucht Drucksachen (Antraege, Gesetzentwuerfe, Kleine/Grosse Anfragen usw.)
     zu Bahn- und Schienenthemen ueber die Titelsuche der DIP-API
  2. Merkt sich, welche Fraktion bzw. welches Organ die Drucksache
     eingebracht hat (Feld "urheber")
  3. Haengt neue Treffer an bahnpolitik.sqlite an
  4. Schreibt aus dem Gesamtbestand die Excel bahnpolitik_gesamt.xlsx

Voraussetzungen:
  - pip: requests und openpyxl
  - API-Key der DIP-API. Entweder als Umgebungsvariable DIP_API_KEY
    (fuer GitHub Actions als Secret) oder in zugangsdaten.txt im selben
    Ordner als Zeile:
        DIP_API_KEY=...
    zugangsdaten.txt gehoert in die .gitignore und niemals ins Repository.
"""

from __future__ import annotations

import os
import re
import sqlite3
import sys
import time as _time
from datetime import date, datetime

import requests

# ---------------------------------------------------------------- Einstellungen

BASE_URL = "https://search.dip.bundestag.de/api/v1"

# Suchbegriffe fuer die Titelsuche. Jeder Begriff wird einzeln abgefragt,
# Dubletten werden ueber die Drucksachen-ID entfernt.
# Bewusst KEIN blankes "Bahn" (sonst matcht auch "Autobahn").
KEYWORDS = [
    "Schienenverkehr",
    "Schienennetz",
    "Schieneninfrastruktur",
    "Eisenbahn",
    "Deutsche Bahn",
    "Bahnstrecke",
    "Bahnhof",
    "Nahverkehr",
    "Regionalisierungsmittel",
    "Deutschlandtakt",
    "Deutschlandticket",
]

# Themen-Verschlagwortung fuer die Auswertung. Eine Drucksache kann mehrere
# Themen tragen. Schluessel = Anzeigename, Werte = Suchmuster im Titel.
THEMEN = {
    "Infrastruktur/Netz": ["schienennetz", "schieneninfrastruktur",
                           "bahnstrecke", "ausbau", "sanierung",
                           "generalsanierung", "elektrifizierung", "gleis"],
    "Bahnhoefe": ["bahnhof", "bahnhöfe", "haltepunkt", "station"],
    "Nahverkehr/SPNV": ["nahverkehr", "regionalisierungsmittel", "spnv",
                        "regionalverkehr"],
    "Deutschlandticket": ["deutschlandticket", "49-euro", "klimaticket"],
    "Deutschlandtakt/Fahrplan": ["deutschlandtakt", "fahrplan", "takt"],
    "Puenktlichkeit/Qualitaet": ["pünktlichkeit", "puenktlichkeit",
                                 "verspätung", "qualität", "ausfall",
                                 "zugausfall", "störung"],
    "Sicherheit": ["sicherheit", "kriminalität", "unfall", "gefährdung",
                   "bundespolizei", "übergriff"],
    "Finanzierung": ["finanzierung", "eigenkapital", "haushalt",
                     "mittel", "milliarden", "investition", "zuschuss"],
    "Gueterverkehr": ["güterverkehr", "gueterverkehr", "schienengüter",
                      "einzelwagen", "kombinierter verkehr"],
    "Konzernstruktur/DB AG": ["deutsche bahn ag", "db ag", "konzern",
                              "infrago", "vorstand", "privatisierung",
                              "strukturreform"],
    "Personal": ["personal", "beschäftigte", "fachkräfte", "lokführer",
                 "streik", "tarif"],
    "Barrierefreiheit": ["barrierefrei", "barrierefreiheit", "mobilitätseingeschränkt"],
}

# Ab wann gesammelt wird. Ueber die Umgebungsvariable DIP_DATE_START
# uebersteuerbar, z.B. DIP_DATE_START=2005-10-18 fuer den gesamten
# Bestand des neuen DIP. Achtung: je frueher, desto laenger laeuft die
# Sammlung und desto groesser wird die Excel.
DATE_START = os.getenv("DIP_DATE_START", "2021-10-26")

# Nur Bundestags-Drucksachen (BT), keine Bundesrats-Drucksachen (BR)
ZUORDNUNG = "BT"

SLEEP_BETWEEN_CALLS = 0.7   # Sekunden, schont das Rate-Limit
MAX_PAGES_PER_QUERY = 200   # Notbremse gegen Endlosschleifen

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "bahnpolitik.sqlite")
XLSX_PATH = os.path.join(BASE_DIR, "bahnpolitik_gesamt.xlsx")
CRED_PATH = os.path.join(BASE_DIR, "zugangsdaten.txt")

SCRIPT_VERSION = 3

# Vorgangspositionen werden nur fuer Drucksachentypen geholt, ueber die
# ueberhaupt abgestimmt werden kann. Kleine Anfragen und Antworten sparen
# wir uns - das spart pro Lauf hunderte API-Aufrufe.
ABSTIMMUNGSFAEHIG = ["Antrag", "Entschließungsantrag", "Änderungsantrag",
                     "Gesetzentwurf", "Beschlussempfehlung und Bericht"]

# Obergrenze, damit ein Lauf in GitHub Actions nicht ausufert.
MAX_VORGAENGE = 400

# Drucksachentypen, bei denen der Urheber wirklich eine eigene politische
# Initiative ergreift. Antworten und Beschlussempfehlungen sind Reaktionen
# und verzerren jeden Fraktionsvergleich.
INITIATIV_TYPEN = ["Kleine Anfrage", "Grosse Anfrage", "Große Anfrage",
                   "Antrag", "Entschließungsantrag", "Gesetzentwurf"]

# Regierungsfraktionen je Wahlperiode (fuer die Spalte "Rolle").
REGIERUNGSFRAKTIONEN = {
    "20": {"SPD", "Gruene", "FDP"},        # Ampel
    "21": {"CDU/CSU", "SPD"},              # ab 2025
}
FRAKTIONEN = ["CDU/CSU", "SPD", "Gruene", "FDP", "AfD", "Linke", "BSW",
              "Mehrere Fraktionen"]

# Parteifarben fuer alle Diagramme. Ohne diese Zuordnung vergibt Excel
# Standardfarben in Reihenfolge der Zeilen - dann ist die AfD gruen und
# die Gruenen sind rot, was bei einer Parteiengrafik zu Fehllesungen fuehrt.
PARTEIFARBEN = {
    "CDU/CSU": "1A1A1A",
    "SPD": "E3000F",
    "Gruene": "1AA037",
    "FDP": "FFCC00",
    "AfD": "009EE0",
    "Linke": "BE3075",
    "BSW": "7D2181",
    "Mehrere Fraktionen": "9E9E9E",
    "Bundesregierung": "5A6B7B",
    "Ausschuesse": "B9BFC6",
}


# ---------------------------------------------------------------- Zugangsdaten

def load_api_key() -> str:
    key = os.getenv("DIP_API_KEY", "")
    if not key and os.path.exists(CRED_PATH):
        with open(CRED_PATH, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if line.startswith("DIP_API_KEY="):
                    key = line.partition("=")[2].strip()
    return key.strip()


# ---------------------------------------------------------------- API

class DipClient:
    def __init__(self, api_key: str):
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"ApiKey {api_key}",
            "Accept": "application/json",
        })

    def drucksachen(self, keyword: str):
        """Liefert alle Drucksachen-Treffer der Titelsuche fuer ein Keyword.
        Folgt der Cursor-Paginierung, bis sich der Cursor nicht mehr aendert."""
        params = {
            "f.titel": keyword,
            "f.zuordnung": ZUORDNUNG,
            "f.datum.start": DATE_START,
        }
        cursor = None
        docs = []
        for _page in range(MAX_PAGES_PER_QUERY):
            if cursor:
                params["cursor"] = cursor
            resp = self.session.get(f"{BASE_URL}/drucksache",
                                    params=params, timeout=30)
            if resp.status_code in (401, 403):
                body = resp.text[:200].replace("\n", " ")
                print(f"\nFEHLER: API verweigert den Zugriff "
                      f"(HTTP {resp.status_code}): {body}")
                print("-> API-Key pruefen.")
                sys.exit(1)
            if resp.status_code == 429:
                _time.sleep(5)
                continue
            resp.raise_for_status()
            data = resp.json()
            docs.extend(data.get("documents", []))
            new_cursor = data.get("cursor")
            _time.sleep(SLEEP_BETWEEN_CALLS)
            if not new_cursor or new_cursor == cursor:
                break
            cursor = new_cursor
        return docs

    def vorgangspositionen(self, vorgang_id: str):
        """Alle Positionen eines Vorgangs. Enthaelt u.a. die
        Beschlussfassung mit dem Abstimmungsergebnis."""
        params = {"f.vorgang": vorgang_id}
        cursor = None
        docs = []
        for _page in range(20):
            if cursor:
                params["cursor"] = cursor
            resp = self.session.get(f"{BASE_URL}/vorgangsposition",
                                    params=params, timeout=30)
            if resp.status_code in (401, 403):
                print(f"FEHLER: Zugriff auf /vorgangsposition verweigert "
                      f"(HTTP {resp.status_code}).")
                return docs
            if resp.status_code == 404:
                return docs
            if resp.status_code == 429:
                _time.sleep(5)
                continue
            resp.raise_for_status()
            data = resp.json()
            docs.extend(data.get("documents", []))
            new_cursor = data.get("cursor")
            _time.sleep(SLEEP_BETWEEN_CALLS)
            if not new_cursor or new_cursor == cursor:
                break
            cursor = new_cursor
        return docs


# ---------------------------------------------------------------- Aufbereitung

_FRAKTIONS_MUSTER = [
    ("cdu/csu", "CDU/CSU"),
    ("bündnis 90", "Gruene"),
    ("grünen", "Gruene"),
    ("fdp", "FDP"),
    ("afd", "AfD"),
    ("bsw", "BSW"),
    ("linke", "Linke"),
    ("spd", "SPD"),
]

_ORGAN_MUSTER = [
    ("bundesregierung", "Bundesregierung"),
    ("bundesrat", "Bundesrat"),
    ("bundesrechnungshof", "Bundesrechnungshof"),
]


def normalize_urheber(raw: str) -> str:
    """Rohes Urheber-Feld auf einen kurzen Fraktions-/Organnamen eindampfen.

    Wichtig: Nennt das Feld mehrere Fraktionen (Gemeinschaftsantraege,
    Wahlvorschlaege), wird NICHT die erste zufaellig getroffene Fraktion
    gewaehlt, sondern "Mehrere Fraktionen" gesetzt. Version 1 hat solche
    Drucksachen still der CDU/CSU zugeschlagen.
    """
    low = (raw or "").lower()
    treffer = {name for needle, name in _FRAKTIONS_MUSTER if needle in low}
    if len(treffer) > 1:
        return "Mehrere Fraktionen"
    if len(treffer) == 1:
        return treffer.pop()
    for needle, name in _ORGAN_MUSTER:
        if needle in low:
            return name
    return (raw or "").strip() or "unbekannt"


def rolle(fraktion: str, wahlperiode: str, datum: str) -> str:
    """Regierungsfraktion, Opposition, Regierung, Ausschuss oder Sonstige."""
    if fraktion == "Bundesregierung":
        return "Regierung"
    if fraktion in ("Bundesrat", "Bundesrechnungshof"):
        return "Sonstige"
    if "ausschuss" in fraktion.lower():
        return "Ausschuss"
    if fraktion not in FRAKTIONS_SET:
        return "Sonstige"
    wp = str(wahlperiode or "").strip()
    if wp not in REGIERUNGSFRAKTIONEN:
        wp = "21" if (datum or "") >= "2025-03-25" else "20"
    if fraktion == "Mehrere Fraktionen":
        return "Fraktionsuebergreifend"
    return ("Regierungsfraktion" if fraktion in REGIERUNGSFRAKTIONEN[wp]
            else "Opposition")


FRAKTIONS_SET = set(FRAKTIONEN)


def themen_aus_titel(titel: str) -> str:
    """Mehrfach-Verschlagwortung. Liefert die Themen als '; '-Liste."""
    low = (titel or "").lower()
    gefunden = [name for name, muster in THEMEN.items()
                if any(m in low for m in muster)]
    return "; ".join(gefunden) if gefunden else "Sonstiges"


_BEZUG_RE = re.compile(r"Drucksachen?\s+((?:\d{2}/\d+)(?:\s*,\s*\d{2}/\d+)*)")


def bezug_aus_titel(titel: str) -> str:
    """Drucksachennummer, auf die sich eine Antwort/Beschlussempfehlung
    bezieht. Bei mehreren Nummern wird die erste genommen."""
    m = _BEZUG_RE.search(titel or "")
    if not m:
        return ""
    return m.group(1).split(",")[0].strip()


def kurztitel(titel: str) -> str:
    """Bei Antworten steht der Sachtitel erst nach dem Bezug in Zeile 3."""
    teile = [t.strip() for t in (titel or "").split("\n") if t.strip()]
    if not teile:
        return ""
    for t in teile:
        if not t.lower().startswith(("auf die", "zu dem", "zu den",
                                     "a) zu dem", "- drucksache")):
            return t
    return teile[-1]


# ------------------------------------------------- Abstimmungen auswerten

# Die DIP-Bemerkung zum Abstimmungsergebnis ist Fliesstext. Typische Formen:
#   "Annahme ... mit den Stimmen der Fraktionen SPD, BUENDNIS 90/DIE GRUENEN
#    und FDP gegen die Stimmen der Fraktionen CDU/CSU und AfD bei Enthaltung
#    der Fraktion DIE LINKE."
#   "Dafuer: CDU/CSU, SPD. Dagegen: AfD. Enthaltung: DIE LINKE."
# Der Parser schneidet den Text in drei Zonen und sucht in jeder Zone nach
# Fraktionsnamen. Was er nicht sicher erkennt, laesst er weg - lieber eine
# Luecke als eine erfundene Stimme.

_ZONEN = [
    ("dafuer", [r"mit den stimmen der fraktion(?:en)?",
                r"\bdaf(?:ü|ue)r\s*:", r"\bzustimmung\s*:",
                r"auf antrag der fraktion(?:en)?"]),
    ("dagegen", [r"gegen die stimmen der fraktion(?:en)?",
                 r"\bgegen die stimmen\b", r"\bdagegen\s*:",
                 r"\bablehnung durch\b"]),
    ("enthaltung", [r"bei (?:stimm)?enthaltung der fraktion(?:en)?",
                    r"bei (?:stimm)?enthaltung",
                    r"\benthaltung(?:en)?\s*:"]),
]

_HALTUNG_LABEL = {"dafuer": "dafuer", "dagegen": "dagegen",
                  "enthaltung": "enthaltung"}


def _fraktionen_in(text: str) -> set:
    low = (text or "").lower()
    return {name for needle, name in _FRAKTIONS_MUSTER if needle in low}


def parse_abstimmung(bemerkung: str) -> dict:
    """Liefert {Fraktion: 'dafuer'|'dagegen'|'enthaltung'}.

    Leeres Dict, wenn der Text kein auswertbares Fraktionsverhalten enthaelt.
    """
    text = (bemerkung or "").strip()
    if not text:
        return {}
    low = text.lower()

    # Alle Zonenanfaenge mit ihrer Position im Text sammeln.
    marker = []
    for haltung, muster in _ZONEN:
        for m in muster:
            for treffer in re.finditer(m, low):
                marker.append((treffer.start(), treffer.end(), haltung))
    if not marker:
        # Einstimmigkeit ohne Fraktionsnennung ist auswertbar, aber ohne
        # Zuordnung zu einzelnen Fraktionen - daher bewusst leer.
        return {}
    marker.sort()

    ergebnis = {}
    for i, (_start, ende, haltung) in enumerate(marker):
        schluss = marker[i + 1][0] if i + 1 < len(marker) else len(text)
        zone = text[ende:schluss]
        # Ein Satzende beendet die Aufzaehlung.
        zone = re.split(r"[.;]", zone)[0]
        for f in _fraktionen_in(zone):
            # Erste Zuordnung gewinnt: "gegen die Stimmen" steht immer nach
            # "mit den Stimmen", eine spaetere Nennung ueberschreibt nicht.
            ergebnis.setdefault(f, _HALTUNG_LABEL[haltung])
    return ergebnis


def beschluesse_aus_position(pos: dict) -> list:
    """Zieht die Beschlussfassungen aus einer Vorgangsposition.

    Die DIP-Feldnamen sind hier defensiv behandelt: die API liefert je nach
    Version leicht abweichende Schluessel. Alles, was nicht gefunden wird,
    bleibt leer - der Rohtext wird in jedem Fall mitgespeichert, damit man
    im Zweifel nachlesen kann.
    """
    roh = pos.get("beschlussfassung") or []
    if isinstance(roh, dict):
        roh = [roh]
    ergebnisse = []
    for b in roh:
        if not isinstance(b, dict):
            continue
        bemerkung = (b.get("abstimm_ergebnis_bemerkung")
                     or b.get("abstimmergebnis_bemerkung")
                     or b.get("bemerkung") or "")
        ergebnisse.append({
            "tenor": b.get("beschlusstenor") or "",
            "abstimmungsart": b.get("abstimmungsart") or "",
            "mehrheit": b.get("mehrheit") or "",
            "grundlage": b.get("grundlage") or "",
            "dokumentnummer": b.get("dokumentnummer") or "",
            "bemerkung": bemerkung,
        })
    return ergebnisse


def rows_from_docs(docs, keyword: str):
    rows = []
    for d in docs:
        urheber_raw = d.get("urheber") or []
        if isinstance(urheber_raw, list):
            urheber_namen = []
            for u in urheber_raw:
                if isinstance(u, dict):
                    urheber_namen.append(u.get("titel") or u.get("bezeichnung") or "")
                else:
                    urheber_namen.append(str(u))
        else:
            urheber_namen = [str(urheber_raw)]
        urheber_join = "; ".join(n for n in urheber_namen if n)
        fraktion = normalize_urheber(urheber_join) if urheber_join else "unbekannt"
        fundstelle = d.get("fundstelle") or {}
        bezug = d.get("vorgangsbezug") or []
        vorgang_ids = []
        for v in bezug:
            if isinstance(v, dict):
                vid = v.get("id") or v.get("vorgangsID") or v.get("vorgang_id")
                if vid:
                    vorgang_ids.append(str(vid))
            elif v:
                vorgang_ids.append(str(v))
        rows.append({
            "id": str(d.get("id", "")),
            "dokumentnummer": d.get("dokumentnummer", ""),
            "typ": d.get("drucksachetyp", ""),
            "datum": d.get("datum", ""),
            "titel": d.get("titel", ""),
            "urheber": urheber_join,
            "fraktion": fraktion,
            "wahlperiode": str(d.get("wahlperiode", "")),
            "pdf_url": fundstelle.get("pdf_url", ""),
            "keyword": keyword,
            "vorgang_ids": ";".join(vorgang_ids),
        })
    return rows


# ---------------------------------------------------------------- SQLite

DDL = """
CREATE TABLE IF NOT EXISTS drucksachen (
    id TEXT PRIMARY KEY,
    dokumentnummer TEXT, typ TEXT, datum TEXT, titel TEXT,
    urheber TEXT, fraktion TEXT, wahlperiode TEXT,
    pdf_url TEXT, keyword TEXT, vorgang_ids TEXT
);

CREATE TABLE IF NOT EXISTS abstimmungen (
    vorgang_id TEXT, position_id TEXT, beschluss_nr INTEGER,
    fraktion TEXT,
    datum TEXT, titel TEXT, vorgangstyp TEXT, drucksache TEXT,
    tenor TEXT, abstimmungsart TEXT, mehrheit TEXT, bemerkung TEXT,
    haltung TEXT,
    PRIMARY KEY (position_id, beschluss_nr, fraktion)
);
"""


def _migriere(con):
    """Bestehende Datenbanken aus Version 1/2 nachruesten."""
    vorhanden = {r[1] for r in con.execute("PRAGMA table_info(drucksachen)")}
    if vorhanden and "vorgang_ids" not in vorhanden:
        con.execute("ALTER TABLE drucksachen ADD COLUMN vorgang_ids TEXT")
        print("Datenbank migriert: Spalte vorgang_ids ergaenzt.")


def upsert(rows) -> int:
    con = sqlite3.connect(DB_PATH)
    con.executescript(DDL)
    _migriere(con)
    n = 0
    for r in rows:
        cur = con.execute(
            """INSERT INTO drucksachen
               (id, dokumentnummer, typ, datum, titel, urheber, fraktion,
                wahlperiode, pdf_url, keyword, vorgang_ids)
               VALUES (:id, :dokumentnummer, :typ, :datum, :titel, :urheber,
                       :fraktion, :wahlperiode, :pdf_url, :keyword,
                       :vorgang_ids)
               ON CONFLICT(id) DO UPDATE SET
                 titel=excluded.titel, urheber=excluded.urheber,
                 fraktion=excluded.fraktion, datum=excluded.datum,
                 vorgang_ids=excluded.vorgang_ids""",
            r)
        n += cur.rowcount
    con.commit()
    con.close()
    return n


def upsert_abstimmungen(zeilen) -> int:
    con = sqlite3.connect(DB_PATH)
    con.executescript(DDL)
    n = 0
    for z in zeilen:
        cur = con.execute(
            """INSERT INTO abstimmungen
               (vorgang_id, position_id, beschluss_nr, fraktion, datum,
                titel, vorgangstyp, drucksache, tenor, abstimmungsart,
                mehrheit, bemerkung, haltung)
               VALUES (:vorgang_id, :position_id, :beschluss_nr, :fraktion,
                       :datum, :titel, :vorgangstyp, :drucksache, :tenor,
                       :abstimmungsart, :mehrheit, :bemerkung, :haltung)
               ON CONFLICT(position_id, beschluss_nr, fraktion)
               DO UPDATE SET
                 haltung=excluded.haltung, tenor=excluded.tenor,
                 bemerkung=excluded.bemerkung, titel=excluded.titel""",
            z)
        n += cur.rowcount
    con.commit()
    con.close()
    return n


def load_all():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.executescript(DDL)
    _migriere(con)
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM drucksachen ORDER BY datum DESC")]
    con.close()
    return rows


def load_abstimmungen():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    con.executescript(DDL)
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM abstimmungen ORDER BY datum DESC, position_id")]
    con.close()
    return rows


# ---------------------------------------------------------------- Excel

BLAU = "1D3C6E"
HELLBLAU = "DCE4F0"
GRAU = "F2F2F2"


def write_excel(rows, abstimmungen=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink

    ARIAL = "Arial"
    h_font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    h_fill = PatternFill("solid", start_color=BLAU)
    thin = Side(style="thin", color="BFBFBF")
    rahmen = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def kopf(ws, werte, row=1):
        for i, v in enumerate(werte, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = h_font
            c.fill = h_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = rahmen
        ws.row_dimensions[row].height = 28

    def breiten(ws, mapping):
        for spalte, w in mapping.items():
            ws.column_dimensions[spalte].width = w

    def arial(ws, min_row=2):
        for row in ws.iter_rows(min_row=min_row):
            for c in row:
                if c.font is None or c.font.name != ARIAL:
                    c.font = Font(name=ARIAL, size=10,
                                  bold=bool(c.font and c.font.bold))

    # ------------------------------------------------ Blatt 1: Drucksachen
    ws = wb.active
    ws.title = "Drucksachen"
    spalten = ["Datum", "Jahr", "Quartal", "Nummer", "Typ", "Urheber",
               "Rolle", "WP", "Themen", "Bezug auf", "Antwort nach (Tage)",
               "Antwort in", "Titel", "Urheber (roh)", "PDF"]
    kopf(ws, spalten)

    sortiert = sorted(rows, key=lambda r: r.get("datum") or "", reverse=True)
    for i, r in enumerate(sortiert, start=2):
        d = r.get("datum") or ""
        try:
            dt = date(int(d[0:4]), int(d[5:7]), int(d[8:10]))
        except (ValueError, IndexError):
            dt = None
        fr = r.get("fraktion", "")
        titel = r.get("titel", "")
        ws.cell(row=i, column=1, value=dt).number_format = "DD.MM.YYYY"
        ws.cell(row=i, column=2, value=f"=IF(A{i}=\"\",\"\",YEAR(A{i}))")
        ws.cell(row=i, column=3,
                value=f"=IF(A{i}=\"\",\"\",YEAR(A{i})&\"-Q\"&ROUNDUP(MONTH(A{i})/3,0))")
        ws.cell(row=i, column=4, value=r.get("dokumentnummer", ""))
        ws.cell(row=i, column=5, value=r.get("typ", ""))
        ws.cell(row=i, column=6, value=fr)
        ws.cell(row=i, column=7,
                value=rolle(fr, r.get("wahlperiode", ""), d))
        ws.cell(row=i, column=8, value=r.get("wahlperiode", ""))
        ws.cell(row=i, column=9, value=themen_aus_titel(titel))
        ws.cell(row=i, column=10, value=bezug_aus_titel(titel))
        ws.cell(row=i, column=11, value=(
            f"=IF(OR(J{i}=\"\",LEFT(E{i},7)<>\"Antwort\"),\"\","
            f"IFERROR(A{i}-INDEX($A$2:$A${len(sortiert)+1},"
            f"MATCH(J{i},$D$2:$D${len(sortiert)+1},0)),\"\"))"))
        ws.cell(row=i, column=12, value=(
            f"=IF(E{i}<>\"Kleine Anfrage\",\"\","
            f"IFERROR(INDEX($D$2:$D${len(sortiert)+1},"
            f"MATCH(D{i},$J$2:$J${len(sortiert)+1},0)),\"offen\"))"))
        ws.cell(row=i, column=13, value=kurztitel(titel))
        ws.cell(row=i, column=14, value=r.get("urheber", ""))
        url = r.get("pdf_url", "")
        c = ws.cell(row=i, column=15, value="PDF" if url else "")
        if url:
            c.hyperlink = Hyperlink(ref=c.coordinate, target=url)
            c.font = Font(name=ARIAL, size=10, color="0563C1", underline="single")

    last = len(sortiert) + 1
    ws.auto_filter.ref = f"A1:O{last}"
    ws.freeze_panes = "E2"
    breiten(ws, {"A": 11, "B": 7, "C": 10, "D": 10, "E": 26, "F": 17,
                 "G": 19, "H": 5, "I": 34, "J": 11, "K": 13, "L": 11,
                 "M": 70, "N": 34, "O": 6})
    for row in ws.iter_rows(min_row=2, max_row=last, min_col=1, max_col=15):
        for c in row:
            if c.font.name != ARIAL:
                c.font = Font(name=ARIAL, size=10)
            c.alignment = Alignment(vertical="top")
    for r_ in range(2, last + 1):
        ws.cell(row=r_, column=11).number_format = "0"

    DS = "Drucksachen"
    RNG = lambda sp: f"{DS}!${sp}$2:${sp}${last}"

    # ------------------------------------------------ Blatt 2: Initiativen
    typen_vorhanden = [t for t in INITIATIV_TYPEN
                       if any(r.get("typ") == t for r in rows)]
    frakt_vorhanden = [f for f in FRAKTIONEN
                       if any(r.get("fraktion") == f for r in rows)]
    # Fraktionen ohne eigene Initiative ans Ende, damit die Torte keine
    # Null-Segmente mit ueberlappender Beschriftung zeigt.
    def _init_zahl(f):
        return sum(1 for r in rows if r.get("fraktion") == f
                   and r.get("typ") in INITIATIV_TYPEN)
    frakt_mit_init = [f for f in frakt_vorhanden if _init_zahl(f) > 0]
    frakt_vorhanden = frakt_mit_init + [f for f in frakt_vorhanden
                                        if f not in frakt_mit_init]

    wi = wb.create_sheet("Initiativen")
    wi["A1"] = "Eigene Initiativen je Fraktion"
    wi["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    wi["A2"] = ("Nur selbst eingebrachte Drucksachen. Antworten der "
                "Bundesregierung und Beschlussempfehlungen der Ausschuesse "
                "sind Reaktionen und hier bewusst nicht enthalten.")
    wi["A2"].font = Font(name=ARIAL, size=9, italic=True)
    wi.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(typen_vorhanden) + 4)

    kopf(wi, ["Fraktion", "Rolle ab WP 21"] + typen_vorhanden
             + ["Gesamt", "Anteil"], row=4)
    for i, f in enumerate(frakt_vorhanden, start=5):
        wi.cell(row=i, column=1, value=f)
        wi.cell(row=i, column=2,
                value=("Regierungsfraktion"
                       if f in REGIERUNGSFRAKTIONEN["21"]
                       else ("Fraktionsuebergreifend"
                             if f == "Mehrere Fraktionen" else "Opposition")))
        for j, t in enumerate(typen_vorhanden, start=3):
            sp = get_column_letter(j)
            wi.cell(row=i, column=j, value=(
                f"=COUNTIFS({RNG('F')},$A{i},{RNG('E')},{sp}$4)"))
        g = get_column_letter(len(typen_vorhanden) + 3)
        a = get_column_letter(3)
        e = get_column_letter(len(typen_vorhanden) + 2)
        wi.cell(row=i, column=len(typen_vorhanden) + 3,
                value=f"=SUM({a}{i}:{e}{i})")
        wi.cell(row=i, column=len(typen_vorhanden) + 4,
                value=f"=IFERROR({g}{i}/{g}${len(frakt_vorhanden)+5},0)")
        wi.cell(row=i, column=len(typen_vorhanden) + 4).number_format = "0.0%"

    sumrow = len(frakt_vorhanden) + 5
    wi.cell(row=sumrow, column=1, value="Gesamt")
    for j in range(3, len(typen_vorhanden) + 4):
        sp = get_column_letter(j)
        wi.cell(row=sumrow, column=j, value=f"=SUM({sp}5:{sp}{sumrow-1})")
    for c in wi[sumrow]:
        c.font = Font(name=ARIAL, bold=True, size=10)
        c.fill = PatternFill("solid", start_color=HELLBLAU)
    breiten(wi, {"A": 20, "B": 20})
    for j in range(3, len(typen_vorhanden) + 5):
        wi.column_dimensions[get_column_letter(j)].width = 16
    arial(wi, min_row=5)

    # ------------------------------------------------ Blatt 3: Zeitverlauf
    quartale = sorted({f"{r['datum'][:4]}-Q{(int(r['datum'][5:7])-1)//3+1}"
                       for r in rows if r.get("datum")})
    wz = wb.create_sheet("Zeitverlauf")
    wz["A1"] = "Drucksachen je Quartal"
    wz["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    kopf(wz, ["Quartal"] + frakt_vorhanden
             + ["Bundesregierung", "Ausschuesse", "Gesamt"], row=3)
    n_f = len(frakt_vorhanden)
    for i, q in enumerate(quartale, start=4):
        wz.cell(row=i, column=1, value=q)
        for j, f in enumerate(frakt_vorhanden, start=2):
            sp = get_column_letter(j)
            wz.cell(row=i, column=j, value=(
                f"=COUNTIFS({RNG('C')},$A{i},{RNG('F')},{sp}$3)"))
        wz.cell(row=i, column=n_f + 2, value=(
            f"=COUNTIFS({RNG('C')},$A{i},{RNG('F')},\"Bundesregierung\")"))
        wz.cell(row=i, column=n_f + 3, value=(
            f"=COUNTIFS({RNG('C')},$A{i},{RNG('G')},\"Ausschuss\")"))
        wz.cell(row=i, column=n_f + 4, value=(
            f"=COUNTIFS({RNG('C')},$A{i})"))
    breiten(wz, {"A": 12})
    for j in range(2, n_f + 5):
        wz.column_dimensions[get_column_letter(j)].width = 15
    arial(wz, min_row=4)

    # ------------------------------------------------ Blatt 4: Themen
    wt = wb.create_sheet("Themen")
    wt["A1"] = "Themen im Zeitverlauf"
    wt["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    wt["A2"] = ("Mehrfachnennung moeglich: eine Drucksache kann mehreren "
                "Themen zugeordnet sein, die Spaltensummen uebersteigen "
                "daher die Zahl der Drucksachen.")
    wt["A2"].font = Font(name=ARIAL, size=9, italic=True)
    jahre = sorted({r["datum"][:4] for r in rows if r.get("datum")})
    kopf(wt, ["Thema"] + jahre + ["Gesamt", "davon Opposition"], row=4)
    themenliste = list(THEMEN.keys()) + ["Sonstiges"]
    # Nach Haeufigkeit sortieren, damit das Diagramm eine Rangfolge zeigt.
    _th_count = {th: sum(1 for r in rows
                         if th in themen_aus_titel(r.get("titel", "")))
                 for th in themenliste}
    themenliste.sort(key=lambda t: -_th_count[t])
    for i, th in enumerate(themenliste, start=5):
        wt.cell(row=i, column=1, value=th)
        for j, jr in enumerate(jahre, start=2):
            sp = get_column_letter(j)
            wt.cell(row=i, column=j, value=(
                f"=SUMPRODUCT(ISNUMBER(SEARCH($A{i},{RNG('I')}))*"
                f"({RNG('B')}={sp}$4))"))
        g = get_column_letter(len(jahre) + 2)
        wt.cell(row=i, column=len(jahre) + 2, value=(
            f"=SUM(B{i}:{get_column_letter(len(jahre)+1)}{i})"))
        wt.cell(row=i, column=len(jahre) + 3, value=(
            f"=SUMPRODUCT(ISNUMBER(SEARCH($A{i},{RNG('I')}))*"
            f"({RNG('G')}=\"Opposition\"))"))
    for j, jr in enumerate(jahre, start=2):
        wt.cell(row=4, column=j).value = int(jr)
    breiten(wt, {"A": 28})
    for j in range(2, len(jahre) + 4):
        wt.column_dimensions[get_column_letter(j)].width = 13
    arial(wt, min_row=5)

    # ------------------------------------------------ Blatt 5: Antwortzeiten
    wa = wb.create_sheet("Antwortzeiten")
    wa["A1"] = "Bearbeitungsdauer Kleiner Anfragen"
    wa["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    wa["A2"] = ("Tage zwischen dem Datum der Kleinen Anfrage und dem Datum "
                "der Antwort der Bundesregierung. Zuordnung ueber die im "
                "Antworttitel genannte Drucksachennummer; Anfragen ohne "
                "auffindbare Antwort bleiben unberuecksichtigt.")
    wa["A2"].font = Font(name=ARIAL, size=9, italic=True)
    wa.merge_cells("A2:F2")
    kopf(wa, ["Fragende Fraktion", "Beantwortete Anfragen",
              "Mittelwert (Tage)", "Kuerzeste", "Laengste",
              "Gestellte Anfragen gesamt"], row=4)

    # Zuordnung Antwort -> anfragende Fraktion, damit COUNTIFS greifen kann.
    nummer_zu_fraktion = {r.get("dokumentnummer"): r.get("fraktion")
                          for r in rows}
    hilfsspalte_start = 16  # Spalte P auf dem Drucksachen-Blatt
    ws.cell(row=1, column=hilfsspalte_start, value="Anfrage von")
    ws.cell(row=1, column=hilfsspalte_start).font = Font(
        name=ARIAL, bold=True, size=9, color="808080")
    for i, r in enumerate(sortiert, start=2):
        bez = bezug_aus_titel(r.get("titel", ""))
        wert = ""
        if str(r.get("typ", "")).startswith("Antwort") and bez:
            wert = nummer_zu_fraktion.get(bez, "")
        ws.cell(row=i, column=hilfsspalte_start, value=wert)
    ws.column_dimensions[get_column_letter(hilfsspalte_start)].width = 14

    for i, f in enumerate(frakt_vorhanden, start=5):
        wa.cell(row=i, column=1, value=f)
        wa.cell(row=i, column=2, value=(
            f"=COUNTIFS({RNG('P')},$A{i},{RNG('K')},\">0\")"))
        wa.cell(row=i, column=3, value=(
            f"=IF($B{i}=0,\"\",IFERROR(AVERAGEIFS({RNG('K')},{RNG('P')},$A{i},"
            f"{RNG('K')},\">0\"),\"\"))"))
        wa.cell(row=i, column=4, value=(
            f"=IF($B{i}=0,\"\",IFERROR(_xlfn.MINIFS({RNG('K')},{RNG('P')},$A{i},"
            f"{RNG('K')},\">0\"),\"\"))"))
        wa.cell(row=i, column=5, value=(
            f"=IF($B{i}=0,\"\",IFERROR(_xlfn.MAXIFS({RNG('K')},{RNG('P')},$A{i},"
            f"{RNG('K')},\">0\"),\"\"))"))
        wa.cell(row=i, column=6, value=(
            f"=COUNTIFS({RNG('F')},$A{i},{RNG('E')},\"Kleine Anfrage\")"))
        for col in (3, 4, 5):
            wa.cell(row=i, column=col).number_format = "0.0"
    breiten(wa, {"A": 22, "B": 20, "C": 17, "D": 12, "E": 12, "F": 22})
    arial(wa, min_row=5)

    # ------------------------------------------------ Blatt 0: Dashboard
    wd = wb.create_sheet("Dashboard", 0)
    wd["A1"] = "Bahnpolitik im Deutschen Bundestag"
    wd["A1"].font = Font(name=ARIAL, bold=True, size=16, color=BLAU)
    wd["A2"] = "Auswertung der DIP-API des Deutschen Bundestages"
    wd["A2"].font = Font(name=ARIAL, size=10, italic=True, color="606060")

    kacheln = [
        ("Drucksachen gesamt", f"=COUNTA({RNG('D')})", "0"),
        ("davon Kleine Anfragen",
         f"=COUNTIF({RNG('E')},\"Kleine Anfrage\")", "0"),
        ("davon Antworten der Regierung",
         f"=COUNTIF({RNG('E')},\"Antwort\")", "0"),
        ("Antraege und Gesetzentwuerfe",
         f"=COUNTIF({RNG('E')},\"Antrag\")+COUNTIF({RNG('E')},\"Gesetzentwurf\")"
         f"+COUNTIF({RNG('E')},\"Entschließungsantrag\")", "0"),
        ("Initiativen der Opposition",
         f"=COUNTIF({RNG('G')},\"Opposition\")", "0"),
        ("Initiativen der Regierungsfraktionen",
         f"=COUNTIF({RNG('G')},\"Regierungsfraktion\")", "0"),
        ("Aktivste Fraktion",
         f"=INDEX(Initiativen!$A$5:$A${sumrow-1},MATCH(MAX(Initiativen!"
         f"${get_column_letter(len(typen_vorhanden)+3)}$5:"
         f"${get_column_letter(len(typen_vorhanden)+3)}${sumrow-1}),"
         f"Initiativen!${get_column_letter(len(typen_vorhanden)+3)}$5:"
         f"${get_column_letter(len(typen_vorhanden)+3)}${sumrow-1},0))", "@"),
        ("Mittlere Antwortzeit (Tage)",
         f"=IFERROR(AVERAGEIFS({RNG('K')},{RNG('K')},\">0\"),\"\")", "0.0"),
        ("Erste Drucksache", f"=MIN({RNG('A')})", "DD.MM.YYYY"),
        ("Letzte Drucksache", f"=MAX({RNG('A')})", "DD.MM.YYYY"),
    ]
    zeile = 4
    for label, formel, fmt in kacheln:
        a = wd.cell(row=zeile, column=1, value=label)
        a.font = Font(name=ARIAL, size=10, color="404040")
        a.fill = PatternFill("solid", start_color=GRAU)
        a.border = rahmen
        b = wd.cell(row=zeile, column=2, value=formel)
        b.font = Font(name=ARIAL, bold=True, size=12, color=BLAU)
        b.number_format = fmt
        b.alignment = Alignment(horizontal="right")
        b.border = rahmen
        zeile += 1

    zeile += 1
    wd.cell(row=zeile, column=1, value="Lesehinweise und Datenqualitaet").font = \
        Font(name=ARIAL, bold=True, size=12, color=BLAU)
    zeile += 1
    hinweise = [
        "Erhebung ueber die Titelsuche der DIP-API. Eine Drucksache wird nur "
        "erfasst, wenn einer der Suchbegriffe im Titel steht - Vollstaendigkeit "
        "ist damit ausdruecklich nicht gegeben.",
        "Suchbegriffe: " + ", ".join(KEYWORDS),
        "Die Spalte 'Rolle ab WP 21' beschreibt den heutigen Status. Aeltere Drucksachen derselben Fraktion koennen aus der Oppositionszeit stammen - massgeblich ist die Spalte 'Rolle' auf dem Blatt Drucksachen.",
        "Der Fraktionsvergleich auf dem Blatt 'Initiativen' zaehlt nur selbst "
        "eingebrachte Drucksachen. Antworten der Bundesregierung werden der "
        "Regierung zugerechnet, nicht der fragenden Fraktion.",
        "Regierungsfraktionen stellen praktisch keine Kleinen Anfragen an die "
        "eigene Regierung. Hohe Zahlen bei Oppositionsfraktionen sind daher "
        "kein Mass fuer bahnpolitisches Engagement, sondern folgen aus der "
        "parlamentarischen Rolle. Die Spalte 'Rolle' macht das sichtbar.",
        "Wahlperiode 20 (bis Maerz 2025): Regierung aus SPD, Gruenen und FDP. "
        "Wahlperiode 21: Regierung aus CDU/CSU und SPD.",
        "Gemeinschafts-Drucksachen mehrerer Fraktionen laufen unter "
        "'Mehrere Fraktionen' und werden keiner einzelnen Fraktion "
        "zugerechnet.",
        "Themen werden aus dem Titel abgeleitet, Mehrfachzuordnung ist "
        "moeglich und beabsichtigt.",
        f"Erstellt am {datetime.now().strftime('%d.%m.%Y %H:%M')} "
        f"durch bahnpolitik_dip.py Version {SCRIPT_VERSION}.",
    ]
    for h in hinweise:
        c = wd.cell(row=zeile, column=1, value="\u2022 " + h)
        c.font = Font(name=ARIAL, size=9, color="404040")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        wd.merge_cells(start_row=zeile, start_column=1,
                       end_row=zeile, end_column=6)
        wd.row_dimensions[zeile].height = 26
        zeile += 1

    wd.column_dimensions["A"].width = 42
    wd.column_dimensions["B"].width = 18
    for sp in "CDEF":
        wd.column_dimensions[sp].width = 14

    # ------------------------------------------------ Diagramme
    from openpyxl.chart import (BarChart, LineChart, PieChart,
                                RadarChart, Reference)
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.marker import DataPoint

    def stil(ch, titel, hoehe=9.5, breite=20):
        ch.title = titel
        ch.height = hoehe
        ch.width = breite
        ch.style = 2
        return ch

    # 1) Initiativen: gestapelte Saeulen, Zusammensetzung je Fraktion
    n_typ = len(typen_vorhanden)
    ch1 = BarChart()
    ch1.type = "col"
    ch1.grouping = "stacked"
    ch1.overlap = 100
    daten = Reference(wi, min_col=3, max_col=n_typ + 2,
                      min_row=4, max_row=sumrow - 1)
    kats = Reference(wi, min_col=1, min_row=5, max_row=sumrow - 1)
    ch1.add_data(daten, titles_from_data=True)
    ch1.set_categories(kats)
    ch1.y_axis.title = "Drucksachen"
    stil(ch1, "Eigene Initiativen je Fraktion, nach Drucksachentyp")
    wi.add_chart(ch1, f"A{sumrow + 3}")

    # 1b) Initiativen: Torte, Anteil der Fraktionen am Gesamtaufkommen
    # Nur Fraktionen mit mindestens einer Initiative, sonst haengen
    # Null-Prozent-Beschriftungen im Titel.
    ende_torte = 4 + len(frakt_mit_init)
    ch1b = PieChart()
    ch1b.add_data(Reference(wi, min_col=n_typ + 3,
                            min_row=4, max_row=ende_torte),
                  titles_from_data=True)
    ch1b.set_categories(Reference(wi, min_col=1, min_row=5,
                                  max_row=ende_torte))
    ch1b.dataLabels = DataLabelList()
    ch1b.dataLabels.showPercent = True
    ch1b.dataLabels.showCatName = True
    ch1b.dataLabels.showVal = True
    ch1b.dataLabels.showSerName = False
    ch1b.dataLabels.showLegendKey = False
    # Jedes Segment in der Farbe seiner Partei.
    for idx, f in enumerate(frakt_mit_init):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = PARTEIFARBEN.get(f, "9E9E9E")
        pt.graphicalProperties.line.solidFill = "FFFFFF"
        ch1b.series[0].data_points.append(pt)
    stil(ch1b, "Anteil der Fraktionen an allen Initiativen",
         hoehe=10, breite=15)
    wi.add_chart(ch1b, f"A{sumrow + 24}")

    # 2) Zeitverlauf: gestapelte Saeulen je Quartal + Linie Gesamt
    letzte_q = len(quartale) + 3
    ch2 = BarChart()
    ch2.type = "col"
    ch2.grouping = "stacked"
    ch2.overlap = 100
    daten = Reference(wz, min_col=2, max_col=n_f + 3,
                      min_row=3, max_row=letzte_q)
    kats = Reference(wz, min_col=1, min_row=4, max_row=letzte_q)
    ch2.add_data(daten, titles_from_data=True)
    ch2.set_categories(kats)
    for serie, name in zip(ch2.series,
                           frakt_vorhanden + ["Bundesregierung", "Ausschuesse"]):
        serie.graphicalProperties.solidFill = PARTEIFARBEN.get(name, "9E9E9E")
    ch2.y_axis.title = "Drucksachen"
    ch2.x_axis.title = "Quartal"
    stil(ch2, "Bahnpolitische Drucksachen je Quartal", breite=26)
    wz.add_chart(ch2, f"A{letzte_q + 3}")

    ch2b = LineChart()
    daten = Reference(wz, min_col=n_f + 4, min_row=3, max_row=letzte_q)
    ch2b.add_data(daten, titles_from_data=True)
    ch2b.set_categories(kats)
    ch2b.y_axis.title = "Drucksachen"
    stil(ch2b, "Gesamtaufkommen im Zeitverlauf", hoehe=7.5, breite=26)
    wz.add_chart(ch2b, f"A{letzte_q + 23}")

    # 3) Themen: waagerechte Balken, Rangfolge
    letzte_th = len(themenliste) + 4
    ch3 = BarChart()
    ch3.type = "bar"
    ch3.grouping = "clustered"
    daten = Reference(wt, min_col=len(jahre) + 2, max_col=len(jahre) + 3,
                      min_row=4, max_row=letzte_th)
    kats = Reference(wt, min_col=1, min_row=5, max_row=letzte_th)
    ch3.add_data(daten, titles_from_data=True)
    ch3.set_categories(kats)
    ch3.x_axis.title = "Nennungen"
    stil(ch3, "Themen insgesamt und Anteil der Opposition", hoehe=11)
    wt.add_chart(ch3, f"A{letzte_th + 3}")

    # 4) Antwortzeiten: Spannweite je Fraktion
    letzte_aw = len(frakt_vorhanden) + 4
    ch4 = BarChart()
    ch4.type = "col"
    ch4.grouping = "clustered"
    daten = Reference(wa, min_col=3, max_col=5, min_row=4, max_row=letzte_aw)
    kats = Reference(wa, min_col=1, min_row=5, max_row=letzte_aw)
    ch4.add_data(daten, titles_from_data=True)
    ch4.set_categories(kats)
    ch4.y_axis.title = "Tage"
    stil(ch4, "Bearbeitungsdauer Kleiner Anfragen je Fraktion")
    wa.add_chart(ch4, f"A{letzte_aw + 3}")

    # 5) Dashboard: Verteilung nach Typ und nach parlamentarischer Rolle
    typen_alle = sorted({r.get("typ") for r in rows if r.get("typ")},
                        key=lambda t: -sum(1 for r in rows if r.get("typ") == t))
    rollen_alle = ["Regierung", "Opposition", "Regierungsfraktion",
                   "Ausschuss", "Fraktionsuebergreifend", "Sonstige"]

    hz = 3
    wd.cell(row=hz, column=14, value="Typ").font = Font(
        name=ARIAL, bold=True, size=9, color="808080")
    wd.cell(row=hz, column=15, value="Anzahl").font = Font(
        name=ARIAL, bold=True, size=9, color="808080")
    for k, t in enumerate(typen_alle, start=hz + 1):
        wd.cell(row=k, column=14, value=t)
        wd.cell(row=k, column=15, value=f"=COUNTIF({RNG('E')},N{k})")
    ende_typ = hz + len(typen_alle)

    hz2 = ende_typ + 2
    wd.cell(row=hz2, column=14, value="Rolle").font = Font(
        name=ARIAL, bold=True, size=9, color="808080")
    wd.cell(row=hz2, column=15, value="Anzahl").font = Font(
        name=ARIAL, bold=True, size=9, color="808080")
    for k, ro in enumerate(rollen_alle, start=hz2 + 1):
        wd.cell(row=k, column=14, value=ro)
        wd.cell(row=k, column=15, value=f"=COUNTIF({RNG('G')},N{k})")
    ende_rolle = hz2 + len(rollen_alle)

    ch5 = PieChart()
    ch5.add_data(Reference(wd, min_col=15, min_row=hz, max_row=ende_typ),
                 titles_from_data=True)
    ch5.set_categories(Reference(wd, min_col=14, min_row=hz + 1,
                                 max_row=ende_typ))
    ch5.dataLabels = DataLabelList()
    ch5.dataLabels.showPercent = True
    stil(ch5, "Drucksachen nach Typ", hoehe=8.5, breite=13)
    wd.add_chart(ch5, f"A{zeile + 2}")

    ch6 = BarChart()
    ch6.type = "bar"
    ch6.add_data(Reference(wd, min_col=15, min_row=hz2, max_row=ende_rolle),
                 titles_from_data=True)
    ch6.set_categories(Reference(wd, min_col=14, min_row=hz2 + 1,
                                 max_row=ende_rolle))
    ch6.legend = None
    stil(ch6, "Drucksachen nach parlamentarischer Rolle",
         hoehe=8.5, breite=13)
    wd.add_chart(ch6, f"A{zeile + 21}")

    # ------------------------------------------------ Blatt 6: Abstimmungen
    HALTUNG_FARBE = {
        "dafuer": ("C6EFCE", "006100", "Dafuer"),
        "dagegen": ("FFC7CE", "9C0006", "Dagegen"),
        "enthaltung": ("FFEB9C", "9C6500", "Enthaltung"),
    }
    wv = wb.create_sheet("Abstimmungen")
    wv["A1"] = "Abstimmungsverhalten der Fraktionen"
    wv["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    wv["A2"] = ("Quelle: Feld 'beschlussfassung' der Vorgangspositionen. "
                "Ueber Kleine Anfragen und deren Antworten wird nicht "
                "abgestimmt - erfasst sind nur Antraege, Gesetzentwuerfe, "
                "Entschliessungsantraege und Beschlussempfehlungen. Leere "
                "Zelle heisst: aus der DIP-Bemerkung nicht eindeutig "
                "ableitbar, nicht etwa Abwesenheit.")
    wv["A2"].font = Font(name=ARIAL, size=9, italic=True)
    wv.merge_cells("A2:J2")
    wv.row_dimensions[2].height = 40

    # Abstimmungen zu Vorlagen buendeln: je (position_id, beschluss_nr)
    # eine Zeile, je Fraktion eine Spalte.
    vorlagen = {}
    for a in (abstimmungen or []):
        schluessel = (a.get("position_id"), a.get("beschluss_nr"))
        eintrag = vorlagen.setdefault(schluessel, {
            "datum": a.get("datum") or "",
            "titel": a.get("titel") or "",
            "drucksache": a.get("drucksache") or "",
            "tenor": a.get("tenor") or "",
            "abstimmungsart": a.get("abstimmungsart") or "",
            "bemerkung": a.get("bemerkung") or "",
            "haltungen": {},
        })
        if a.get("fraktion") and a.get("haltung"):
            eintrag["haltungen"][a["fraktion"]] = a["haltung"]

    liste = sorted(vorlagen.values(), key=lambda v: v["datum"], reverse=True)
    frakt_abst = [f for f in FRAKTIONEN
                  if any(f in v["haltungen"] for v in liste)] or FRAKTIONEN[:6]

    kopf(wv, ["Datum", "Drucksache", "Vorlage", "Beschluss",
              "Abstimmungsart"] + frakt_abst, row=4)

    if not liste:
        wv.cell(row=5, column=1,
                value="Noch keine Beschlussfassungen im Bestand.")
        wv.cell(row=5, column=1).font = Font(name=ARIAL, size=10,
                                             italic=True, color="9C0006")
        wv.cell(row=6, column=1, value=(
            "Das ist ein moegliches Ergebnis, kein Fehler: nicht jede "
            "Vorlage kommt zur Abstimmung, und nicht zu jeder Abstimmung "
            "nennt DIP das Fraktionsverhalten im Klartext."))
        wv.cell(row=6, column=1).font = Font(name=ARIAL, size=9,
                                             color="606060")
        letzte_ab = 6
    else:
        for i, v in enumerate(liste, start=5):
            try:
                dt = date(int(v["datum"][0:4]), int(v["datum"][5:7]),
                          int(v["datum"][8:10]))
            except (ValueError, IndexError, TypeError):
                dt = None
            wv.cell(row=i, column=1, value=dt).number_format = "DD.MM.YYYY"
            wv.cell(row=i, column=2, value=v["drucksache"])
            wv.cell(row=i, column=3, value=v["titel"][:180])
            wv.cell(row=i, column=4, value=v["tenor"])
            wv.cell(row=i, column=5, value=v["abstimmungsart"])
            for j, f in enumerate(frakt_abst, start=6):
                h = v["haltungen"].get(f)
                c = wv.cell(row=i, column=j)
                if h in HALTUNG_FARBE:
                    fill, farbe, label = HALTUNG_FARBE[h]
                    c.value = label
                    c.fill = PatternFill("solid", start_color=fill)
                    c.font = Font(name=ARIAL, size=9, bold=True, color=farbe)
                    c.alignment = Alignment(horizontal="center")
                c.border = rahmen
        letzte_ab = len(liste) + 4
        wv.auto_filter.ref = f"A4:{get_column_letter(5+len(frakt_abst))}{letzte_ab}"
        wv.freeze_panes = "F5"

    breiten(wv, {"A": 11, "B": 12, "C": 62, "D": 26, "E": 20})
    for j in range(6, 6 + len(frakt_abst)):
        wv.column_dimensions[get_column_letter(j)].width = 13

    # Bilanz je Fraktion, per Formel aus der Matrix darueber
    bil = letzte_ab + 3
    wv.cell(row=bil, column=1, value="Bilanz je Fraktion").font = Font(
        name=ARIAL, bold=True, size=12, color=BLAU)
    kopf(wv, ["Fraktion", "Dafuer", "Dagegen", "Enthaltung",
              "Abstimmungen gesamt"], row=bil + 1)
    for k, f in enumerate(frakt_abst, start=bil + 2):
        sp = get_column_letter(6 + frakt_abst.index(f))
        bereich = f"${sp}$5:${sp}${max(letzte_ab, 5)}"
        wv.cell(row=k, column=1, value=f)
        wv.cell(row=k, column=2, value=f"=COUNTIF({bereich},\"Dafuer\")")
        wv.cell(row=k, column=3, value=f"=COUNTIF({bereich},\"Dagegen\")")
        wv.cell(row=k, column=4, value=f"=COUNTIF({bereich},\"Enthaltung\")")
        wv.cell(row=k, column=5, value=f"=SUM(B{k}:D{k})")
    letzte_bil = bil + 1 + len(frakt_abst)
    arial(wv, min_row=bil + 2)

    if liste:
        ch7 = BarChart()
        ch7.type = "bar"
        ch7.grouping = "stacked"
        ch7.overlap = 100
        ch7.add_data(Reference(wv, min_col=2, max_col=4,
                               min_row=bil + 1, max_row=letzte_bil),
                     titles_from_data=True)
        ch7.set_categories(Reference(wv, min_col=1, min_row=bil + 2,
                                     max_row=letzte_bil))
        for serie, farbe in zip(ch7.series, ["4CAF50", "D32F2F", "FBC02D"]):
            serie.graphicalProperties.solidFill = farbe
        ch7.x_axis.title = "Abstimmungen"
        stil(ch7, "Abstimmungsverhalten je Fraktion", hoehe=9, breite=18)
        wv.add_chart(ch7, f"A{letzte_bil + 3}")

    # ------------------------------------------------ Blatt 7: Profil
    wp_ = wb.create_sheet("Profil")
    wp_["A1"] = "Fraktionsprofil: vier messbare Dimensionen"
    wp_["A1"].font = Font(name=ARIAL, bold=True, size=13, color=BLAU)
    wp_["A2"] = ("Achtung Interpretation: Das sind Aktivitaetsmasse, kein "
                 "Qualitaetsurteil. Ob eine Anfrage sachkundig oder blosse "
                 "Symbolpolitik ist, steht in DIP nicht drin und laesst sich "
                 "aus Metadaten nicht ableiten. Ausserdem koennen "
                 "Regierungsfraktionen strukturell kaum Kleine Anfragen "
                 "stellen - niedrige Werte heissen dort nicht Untaetigkeit.")
    wp_["A2"].font = Font(name=ARIAL, size=9, italic=True, color="9C0006")
    wp_.merge_cells("A2:H2")
    wp_.row_dimensions[2].height = 48

    profil_frakt = frakt_mit_init or frakt_vorhanden
    kopf(wp_, ["Kennzahl"] + profil_frakt + ["Was sie misst"], row=4)

    absti = ["Antrag", "Entschließungsantrag", "Gesetzentwurf"]

    # Hilfsraster rechts: Thema x Fraktion und Quartal x Fraktion, je 0/1.
    hspalte = len(profil_frakt) + 4
    hs = get_column_letter(hspalte)
    wp_.cell(row=4, column=hspalte, value="Hilfsraster Themen").font = Font(
        name=ARIAL, bold=True, size=8, color="A0A0A0")
    for ti, th in enumerate(themenliste, start=5):
        wp_.cell(row=ti, column=hspalte, value=th)
        for fj, f in enumerate(profil_frakt, start=hspalte + 1):
            sp = get_column_letter(fj)
            wp_.cell(row=ti, column=fj, value=(
                f"=IF(SUMPRODUCT(ISNUMBER(SEARCH(${hs}{ti},{RNG('I')}))*"
                f"({RNG('F')}={sp}$4))>0,1,0)"))
    th_ende = len(themenliste) + 4

    q_start = th_ende + 2
    wp_.cell(row=q_start, column=hspalte,
             value="Hilfsraster Quartale").font = Font(
        name=ARIAL, bold=True, size=8, color="A0A0A0")
    for qi, q in enumerate(quartale, start=q_start + 1):
        wp_.cell(row=qi, column=hspalte, value=q)
        for fj, f in enumerate(profil_frakt, start=hspalte + 1):
            sp = get_column_letter(fj)
            wp_.cell(row=qi, column=fj, value=(
                f"=IF(COUNTIFS({RNG('C')},${hs}{qi},{RNG('F')},{sp}$4)>0,1,0)"))
    q_ende = q_start + len(quartale)

    # Kopfzeile des Hilfsrasters mit den Fraktionsnamen
    for fj, f in enumerate(profil_frakt, start=hspalte + 1):
        wp_.cell(row=4, column=fj, value=f).font = Font(
            name=ARIAL, bold=True, size=8, color="A0A0A0")

    kennzahlen = [
        ("Initiativen gesamt",
         lambda sp, i: "+".join(
             f"COUNTIFS({RNG('F')},{sp}$4,{RNG('E')},\"{t}\")"
             for t in typen_vorhanden),
         "Wie viel eine Fraktion ueberhaupt einbringt"),
        ("davon abstimmungsfaehig",
         lambda sp, i: "+".join(
             f"COUNTIFS({RNG('F')},{sp}$4,{RNG('E')},\"{t}\")"
             for t in absti if t in typen_vorhanden) or "0",
         "Gestaltung statt blosser Kontrolle: Antraege und Gesetzentwuerfe"),
        ("Themenbreite",
         lambda sp, i: f"SUM({sp}$5:{sp}${th_ende})",
         f"Wie viele der {len(themenliste)} Sachthemen beruehrt werden"),
        ("Aktive Quartale",
         lambda sp, i: f"SUM({sp}${q_start+1}:{sp}${q_ende})",
         f"Beharrlichkeit: Quartale mit Aktivitaet, von {len(quartale)}"),
    ]

    for i, (name, formel, erklaerung) in enumerate(kennzahlen, start=5):
        wp_.cell(row=i, column=1, value=name).font = Font(
            name=ARIAL, bold=True, size=10)
        for j, f in enumerate(profil_frakt, start=2):
            sp = get_column_letter(hspalte + j - 1)
            wp_.cell(row=i, column=j, value="=" + formel(sp, i))
        c = wp_.cell(row=i, column=len(profil_frakt) + 2, value=erklaerung)
        c.font = Font(name=ARIAL, size=9, italic=True, color="606060")

    # Normierte Werte fuer das Netzdiagramm: je Kennzahl auf den
    # Spitzenreiter bezogen, damit die Achsen vergleichbar sind.
    norm = 11
    wp_.cell(row=norm, column=1,
             value="Normiert (Spitzenreiter je Zeile = 100)").font = Font(
        name=ARIAL, bold=True, size=11, color=BLAU)
    kopf(wp_, ["Kennzahl"] + profil_frakt, row=norm + 1)
    letzte_f = get_column_letter(len(profil_frakt) + 1)
    for i in range(4):
        quelle = 5 + i
        ziel = norm + 2 + i
        wp_.cell(row=ziel, column=1, value=kennzahlen[i][0])
        for j in range(2, len(profil_frakt) + 2):
            sp = get_column_letter(j)
            wp_.cell(row=ziel, column=j, value=(
                f"=IFERROR(ROUND({sp}{quelle}/MAX($B${quelle}:${letzte_f}"
                f"${quelle})*100,0),0)"))
    norm_ende = norm + 5

    breiten(wp_, {"A": 26})
    for j in range(2, len(profil_frakt) + 2):
        wp_.column_dimensions[get_column_letter(j)].width = 13
    wp_.column_dimensions[get_column_letter(len(profil_frakt) + 2)].width = 52
    for j in range(hspalte, hspalte + len(profil_frakt) + 1):
        wp_.column_dimensions[get_column_letter(j)].hidden = True
    arial(wp_, min_row=5)

    ch8 = RadarChart()
    ch8.type = "marker"
    ch8.add_data(Reference(wp_, min_col=2, max_col=len(profil_frakt) + 1,
                           min_row=norm + 1, max_row=norm_ende),
                 titles_from_data=True)
    ch8.set_categories(Reference(wp_, min_col=1, min_row=norm + 2,
                                 max_row=norm_ende))
    for serie, name in zip(ch8.series, profil_frakt):
        serie.graphicalProperties.line.solidFill = PARTEIFARBEN.get(
            name, "9E9E9E")
        serie.graphicalProperties.line.width = 22000
    stil(ch8, "Fraktionsprofil im Vergleich (100 = Spitzenwert)",
         hoehe=11, breite=16)
    wp_.add_chart(ch8, f"A{norm_ende + 3}")

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(XLSX_PATH)


# ---------------------------------------------------------------- Sondierung

def sondiere(client) -> int:
    """Prueft empirisch, wie weit der Bestand zurueckreicht.

    Fragt je Wahlperiode und Suchbegriff nur die Trefferzahl ab (numFound),
    laedt also keine Dokumente. Damit sieht man ohne Ratespiel, ab welcher
    Wahlperiode die API ueberhaupt Daten liefert.
    """
    print("\nSondierung: Treffer je Wahlperiode (nur Zaehlung, kein Abruf)")
    print("-" * 60)
    gesamt = {}
    for wp in range(12, 22):
        summe = 0
        fehler = False
        for kw in KEYWORDS:
            try:
                resp = client.session.get(
                    f"{BASE_URL}/drucksache",
                    params={"f.titel": kw, "f.zuordnung": ZUORDNUNG,
                            "f.wahlperiode": wp},
                    timeout=30)
                if resp.status_code != 200:
                    fehler = True
                    break
                summe += resp.json().get("numFound", 0)
            except Exception:
                fehler = True
                break
            _time.sleep(SLEEP_BETWEEN_CALLS)
        gesamt[wp] = None if fehler else summe
        anzeige = "Abruf fehlgeschlagen" if fehler else f"{summe} Treffer"
        print(f"  Wahlperiode {wp}: {anzeige}")
    vorhanden = [wp for wp, v in gesamt.items() if v]
    if vorhanden:
        print(f"\nDaten vorhanden ab Wahlperiode {min(vorhanden)}.")
        print("Zum Sammeln DIP_DATE_START entsprechend setzen, z.B.:")
        print("  DIP_DATE_START=2005-10-18 python bahnpolitik_dip.py")
    else:
        print("\nKeine Treffer - Suchbegriffe oder API-Key pruefen.")
    return 0


# ---------------------------------------------------------------- Ablauf

def main() -> int:
    print("=" * 50)
    print(f" Bahnpolitik-Tracker (DIP-API, Version {SCRIPT_VERSION})")
    print("=" * 50)

    key = load_api_key()
    if not key:
        print(f"\nFEHLER: Kein API-Key gefunden (DIP_API_KEY oder {CRED_PATH}).")
        return 2
    print(f"API-Key geladen: {len(key)} Zeichen")

    client = DipClient(key)

    if "--sondierung" in sys.argv:
        return sondiere(client)

    print(f"Sammle ab {DATE_START}.")
    gefunden = {}
    for kw in KEYWORDS:
        docs = client.drucksachen(kw)
        rows = rows_from_docs(docs, kw)
        neu = 0
        for r in rows:
            if r["id"] and r["id"] not in gefunden:
                gefunden[r["id"]] = r
                neu += 1
        print(f"[{kw}] {len(rows)} Treffer, davon {neu} neu in diesem Lauf")

    n = upsert(list(gefunden.values()))
    print(f"\n{len(gefunden)} eindeutige Drucksachen, {n} gespeichert/aktualisiert.")

    # ---- Abstimmungen: Vorgaenge der abstimmungsfaehigen Drucksachen holen
    vorgang_ids = []
    for r in gefunden.values():
        if r.get("typ") not in ABSTIMMUNGSFAEHIG:
            continue
        for vid in (r.get("vorgang_ids") or "").split(";"):
            if vid and vid not in vorgang_ids:
                vorgang_ids.append(vid)
    vorgang_ids = vorgang_ids[:MAX_VORGAENGE]

    print(f"\n{len(vorgang_ids)} Vorgaenge zu abstimmungsfaehigen Drucksachen.")
    if not vorgang_ids:
        print("Hinweis: Kein Vorgangsbezug geliefert. Entweder enthaelt der "
              "Bestand keine abstimmungsfaehigen Drucksachen, oder das Feld "
              "'vorgangsbezug' heisst in dieser API-Version anders - dann "
              "bitte im Swagger nachsehen.")

    abst_zeilen = []
    ohne_bemerkung = 0
    unparsbar = []
    for nr, vid in enumerate(vorgang_ids, start=1):
        try:
            positionen = client.vorgangspositionen(vid)
        except Exception as exc:
            print(f"  Vorgang {vid}: Abruf fehlgeschlagen ({exc})")
            continue
        for pos in positionen:
            for idx, b in enumerate(beschluesse_aus_position(pos)):
                if not b["bemerkung"]:
                    ohne_bemerkung += 1
                haltungen = parse_abstimmung(b["bemerkung"])
                if b["bemerkung"] and not haltungen:
                    unparsbar.append(b["bemerkung"][:120])
                basis = {
                    "vorgang_id": vid,
                    "position_id": str(pos.get("id", f"{vid}-{idx}")),
                    "beschluss_nr": idx,
                    "datum": pos.get("datum", ""),
                    "titel": pos.get("titel", "") or pos.get("vorgangstyp", ""),
                    "vorgangstyp": pos.get("vorgangstyp", ""),
                    "drucksache": b["dokumentnummer"],
                    "tenor": b["tenor"],
                    "abstimmungsart": b["abstimmungsart"],
                    "mehrheit": b["mehrheit"],
                    "bemerkung": b["bemerkung"],
                }
                if haltungen:
                    for f, h in haltungen.items():
                        z = dict(basis)
                        z["fraktion"] = f
                        z["haltung"] = h
                        abst_zeilen.append(z)
                else:
                    # Beschluss ohne auswertbares Fraktionsverhalten trotzdem
                    # festhalten, damit die Bemerkung nachlesbar bleibt.
                    z = dict(basis)
                    z["fraktion"] = ""
                    z["haltung"] = ""
                    abst_zeilen.append(z)
        if nr % 25 == 0:
            print(f"  {nr}/{len(vorgang_ids)} Vorgaenge abgefragt")

    if abst_zeilen:
        m = upsert_abstimmungen(abst_zeilen)
        gewertet = len({(z["position_id"], z["beschluss_nr"])
                        for z in abst_zeilen if z["fraktion"]})
        print(f"{len(abst_zeilen)} Abstimmungszeilen gespeichert ({m} neu/geaendert), "
              f"{gewertet} Beschluesse mit erkanntem Fraktionsverhalten.")
        if ohne_bemerkung:
            print(f"{ohne_bemerkung} Beschluesse ohne Ergebnis-Bemerkung.")
        if unparsbar:
            print(f"{len(unparsbar)} Bemerkungen nicht auswertbar, Beispiele:")
            for b in unparsbar[:3]:
                print(f"   > {b}")
    else:
        print("Keine Beschlussfassungen gefunden.")

    rows = load_all()
    if not rows:
        print("Kein Bestand vorhanden, keine Excel erzeugt.")
        return 1
    write_excel(rows, load_abstimmungen())

    counts = {}
    for r in rows:
        counts[r["fraktion"]] = counts.get(r["fraktion"], 0) + 1
    print("\nBestand je Fraktion/Organ:")
    for f, c in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  {f}: {c}")
    print(f"\nExcel: {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        print(f"\nUNERWARTETER FEHLER: {type(exc).__name__}: {exc}")
        sys.exit(1)
